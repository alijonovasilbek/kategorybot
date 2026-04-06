from fastapi import FastAPI, HTTPException, File, UploadFile, Form, Header, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta
from PIL import Image as PILImage
import json, os, shutil, uuid, secrets, io, math
import sys
sys.path.append(os.path.dirname(__file__))
from database.crud import *
from database.db import init_db
from bot.config import ADMIN_USERNAME, ADMIN_PASSWORD, SHOP_LAT, SHOP_LNG, DELIVERY_PRICE_PER_KM, DELIVERY_FREE_DISTANCE_KM, ORDER_GROUP_ID

app = FastAPI(title="KategoryBot API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

@app.on_event("startup")
async def startup():
    await init_db()

# ── STATIC FILES ──────────────────────────────────────────────────────────────
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

try:
    app.mount("/", StaticFiles(directory="webapp", html=True), name="webapp")
except: pass

# ── HELPERS ───────────────────────────────────────────────────────────────────
def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

def calc_delivery(lat, lng):
    if not lat or not lng: return 0, 0
    dist = haversine(SHOP_LAT, SHOP_LNG, lat, lng)
    if dist <= DELIVERY_FREE_DISTANCE_KM: return 0, round(dist, 2)
    return round((dist - DELIVERY_FREE_DISTANCE_KM) * DELIVERY_PRICE_PER_KM), round(dist, 2)

def save_image(file, max_size=(800, 800)):
    ext = (file.filename.split(".")[-1] or "jpg").lower()
    if ext not in ("jpg", "jpeg", "png", "webp"): ext = "jpg"
    filename = f"{uuid.uuid4()}.{ext}"
    path = os.path.join(UPLOAD_DIR, filename)
    try:
        img = PILImage.open(file.file)
        img.thumbnail(max_size, PILImage.LANCZOS)
        if img.mode in ("RGBA", "P"): img = img.convert("RGB")
        img.save(path, optimize=True, quality=85)
    except:
        file.file.seek(0)
        with open(path, "wb") as f: shutil.copyfileobj(file.file, f)
    return f"/uploads/{filename}"

async def verify_admin(authorization: str = Header(None)):
    if not authorization or not authorization.startswith("Bearer "): raise HTTPException(401, "Unauthorized")
    token = authorization[7:]
    if not await check_admin_session(token): raise HTTPException(401, "Session expired")
    return token

# ── ADMIN AUTH ────────────────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str

@app.post("/api/admin/login")
async def admin_login(req: LoginRequest):
    if req.username != ADMIN_USERNAME or req.password != ADMIN_PASSWORD:
        raise HTTPException(401, "Noto'g'ri login yoki parol")
    token = secrets.token_urlsafe(32)
    expires = datetime.utcnow() + timedelta(hours=24)
    await create_admin_session(token, expires)
    return {"token": token, "expires_at": expires.isoformat()}

@app.post("/api/admin/logout")
async def admin_logout(token: str = Depends(verify_admin)):
    await delete_admin_session(token)
    return {"ok": True}

@app.get("/api/admin/check")
async def admin_check(token: str = Depends(verify_admin)):
    return {"ok": True}

# ── UPLOAD ────────────────────────────────────────────────────────────────────
@app.post("/api/upload")
async def upload_image(file: UploadFile = File(...)):
    url = save_image(file)
    return {"url": url}

# ── DELIVERY CALC ─────────────────────────────────────────────────────────────
@app.get("/api/delivery/calc")
async def delivery_calc(lat: float, lng: float):
    price, dist = calc_delivery(lat, lng)
    return {"price": price, "distance_km": dist, "free_distance": DELIVERY_FREE_DISTANCE_KM}

# ── USERS ─────────────────────────────────────────────────────────────────────
@app.get("/api/users/{telegram_id}")
async def get_user_api(telegram_id: int):
    user = await get_user(telegram_id)
    if not user: raise HTTPException(404, "User not found")
    return _user_dict(user)

@app.put("/api/users/{telegram_id}/language")
async def set_language(telegram_id: int, lang: str):
    await update_user_language(telegram_id, lang)
    return {"ok": True}

@app.get("/api/admin/users")
async def admin_users(_=Depends(verify_admin)):
    return [_user_dict(u) for u in await get_all_users()]

# ── CATEGORIES ────────────────────────────────────────────────────────────────
@app.get("/api/categories")
async def list_categories(parent_id: Optional[int] = None):
    return [_cat_dict(c) for c in await get_categories(parent_id)]

@app.get("/api/categories/all")
async def list_all_categories():
    return [_cat_dict(c) for c in await get_all_categories()]

@app.post("/api/categories")
async def create_category_api(
    name_uz: str = Form(...), name_ru: str = Form(...),
    parent_id: Optional[int] = Form(None),
    image: Optional[UploadFile] = File(None),
    _=Depends(verify_admin)
):
    image_url = save_image(image) if image and image.filename else None
    cat = await create_category(name_uz, name_ru, image_url, parent_id)
    return _cat_dict(cat)

@app.put("/api/categories/{cat_id}")
async def update_category_api(
    cat_id: int, name_uz: str = Form(...), name_ru: str = Form(...),
    image: Optional[UploadFile] = File(None), _=Depends(verify_admin)
):
    data = {"name_uz": name_uz, "name_ru": name_ru}
    if image and image.filename: data["image_url"] = save_image(image)
    await update_category(cat_id, **data)
    return {"ok": True}

@app.delete("/api/categories/{cat_id}")
async def delete_category_api(cat_id: int, _=Depends(verify_admin)):
    await delete_category(cat_id); return {"ok": True}

# ── PRODUCTS ──────────────────────────────────────────────────────────────────
@app.get("/api/products")
async def list_products(category_id: Optional[int] = None, search: Optional[str] = None, limit: int = 50, offset: int = 0):
    return [_product_dict(p) for p in await get_products(category_id, search, limit, offset)]

@app.get("/api/products/all")
async def list_all_products_api(_=Depends(verify_admin)):
    return [_product_dict(p) for p in await get_all_products()]

@app.get("/api/products/{product_id}")
async def get_product_api(product_id: int):
    p = await get_product(product_id)
    if not p: raise HTTPException(404, "Not found")
    rating, count = await get_product_rating(product_id)
    d = _product_dict(p)
    d["rating"] = rating; d["review_count"] = count
    return d

@app.get("/api/products/{product_id}/similar")
async def similar_products(product_id: int, category_id: int):
    return [_product_dict(p) for p in await get_similar_products(product_id, category_id)]

@app.post("/api/products")
async def create_product_api(
    title_uz: str = Form(...), title_ru: str = Form(...),
    description_uz: str = Form(""), description_ru: str = Form(""),
    price: float = Form(...), old_price: Optional[float] = Form(None),
    category_id: int = Form(...), stock: int = Form(100),
    images: List[UploadFile] = File(default=[]),
    _=Depends(verify_admin)
):
    image_urls = [save_image(img) for img in images[:4] if img.filename]
    p = await create_product(dict(title_uz=title_uz, title_ru=title_ru,
        description_uz=description_uz, description_ru=description_ru,
        price=price, old_price=old_price, category_id=category_id,
        stock=stock, images=image_urls))
    return {"id": p.id}

@app.put("/api/products/{product_id}")
async def update_product_api(
    product_id: int,
    title_uz: str = Form(...), title_ru: str = Form(...),
    description_uz: str = Form(""), description_ru: str = Form(""),
    price: float = Form(...), old_price: Optional[float] = Form(None),
    category_id: int = Form(...), stock: int = Form(100),
    images: List[UploadFile] = File(default=[]),
    existing_images: str = Form("[]"),
    _=Depends(verify_admin)
):
    existing = json.loads(existing_images)
    new_urls = [save_image(img) for img in images if img.filename]
    all_images = (existing + new_urls)[:4]
    await update_product(product_id, title_uz=title_uz, title_ru=title_ru,
        description_uz=description_uz, description_ru=description_ru,
        price=price, old_price=old_price, category_id=category_id,
        stock=stock, images=all_images)
    return {"ok": True}

@app.delete("/api/products/{product_id}")
async def delete_product_api(product_id: int, _=Depends(verify_admin)):
    await delete_product(product_id); return {"ok": True}

# ── REVIEWS ───────────────────────────────────────────────────────────────────
class ReviewRequest(BaseModel):
    user_id: int
    product_id: int
    rating: int
    comment: Optional[str] = None

@app.get("/api/products/{product_id}/reviews")
async def product_reviews(product_id: int):
    reviews = await get_product_reviews(product_id)
    return [{"id": r.id, "rating": r.rating, "comment": r.comment,
             "user_name": r.user.full_name if r.user else "?",
             "created_at": str(r.created_at)} for r in reviews]

@app.post("/api/reviews")
async def post_review(req: ReviewRequest):
    if not 1 <= req.rating <= 5: raise HTTPException(400, "Rating 1-5 orasida bo'lishi kerak")
    await add_review(req.user_id, req.product_id, req.rating, req.comment)
    return {"ok": True}

@app.delete("/api/admin/reviews/{review_id}")
async def delete_review_api(review_id: int, _=Depends(verify_admin)):
    await delete_review(review_id); return {"ok": True}

@app.get("/api/admin/reviews")
async def admin_reviews(_=Depends(verify_admin)):
    reviews = await get_all_reviews()
    return [{"id": r.id, "rating": r.rating, "comment": r.comment,
             "user_name": r.user.full_name if r.user else "?",
             "product_title": r.product.title_uz if r.product else "?",
             "created_at": str(r.created_at)} for r in reviews]

# ── WISHLIST ──────────────────────────────────────────────────────────────────
@app.get("/api/wishlist/{user_id}")
async def get_wishlist_api(user_id: int):
    items = await get_wishlist(user_id)
    return [{"product_id": i.product_id, "product": _product_dict(i.product)} for i in items if i.product]

@app.post("/api/wishlist/{user_id}/{product_id}")
async def toggle_wishlist_api(user_id: int, product_id: int):
    added = await toggle_wishlist(user_id, product_id)
    return {"added": added}

@app.get("/api/wishlist/{user_id}/{product_id}/check")
async def check_wishlist(user_id: int, product_id: int):
    return {"in_wishlist": await is_in_wishlist(user_id, product_id)}

# ── COUPONS ───────────────────────────────────────────────────────────────────
@app.get("/api/coupons/check/{code}")
async def check_coupon(code: str):
    c = await get_coupon(code)
    if not c: raise HTTPException(404, "Kupon topilmadi yoki muddati tugagan")
    return {"id": c.id, "code": c.code, "discount_type": c.discount_type,
            "discount_value": c.discount_value, "min_order_amount": c.min_order_amount}

@app.get("/api/admin/coupons")
async def admin_coupons(_=Depends(verify_admin)):
    coupons = await get_all_coupons()
    return [{"id": c.id, "code": c.code, "discount_type": c.discount_type,
             "discount_value": c.discount_value, "min_order_amount": c.min_order_amount,
             "max_uses": c.max_uses, "used_count": c.used_count,
             "is_active": c.is_active, "expires_at": str(c.expires_at) if c.expires_at else None} for c in coupons]

class CouponRequest(BaseModel):
    code: str
    discount_type: str = "percent"
    discount_value: float
    min_order_amount: float = 0
    max_uses: Optional[int] = None
    expires_at: Optional[str] = None

@app.post("/api/admin/coupons")
async def create_coupon_api(req: CouponRequest, _=Depends(verify_admin)):
    expires = datetime.fromisoformat(req.expires_at) if req.expires_at else None
    c = await create_coupon(req.code, req.discount_type, req.discount_value,
                            req.min_order_amount, req.max_uses, expires)
    return {"id": c.id}

@app.put("/api/admin/coupons/{coupon_id}")
async def update_coupon_api(coupon_id: int, req: CouponRequest, _=Depends(verify_admin)):
    data = req.dict(exclude_none=True)
    if "expires_at" in data and data["expires_at"]:
        data["expires_at"] = datetime.fromisoformat(data["expires_at"])
    await update_coupon(coupon_id, **data)
    return {"ok": True}

@app.delete("/api/admin/coupons/{coupon_id}")
async def delete_coupon_api(coupon_id: int, _=Depends(verify_admin)):
    await delete_coupon(coupon_id); return {"ok": True}

# ── CART ──────────────────────────────────────────────────────────────────────
class CartRequest(BaseModel):
    user_id: int; product_id: int; quantity: int = 1

@app.get("/api/cart/{user_id}")
async def get_cart_api(user_id: int):
    items = await get_cart(user_id)
    return [{"id": i.id, "product_id": i.product_id, "quantity": i.quantity, "product": _product_dict(i.product)} for i in items]

@app.post("/api/cart")
async def add_cart(req: CartRequest):
    await add_to_cart(req.user_id, req.product_id, req.quantity); return {"ok": True}

@app.put("/api/cart")
async def update_cart(req: CartRequest):
    await update_cart_item(req.user_id, req.product_id, req.quantity); return {"ok": True}

@app.delete("/api/cart/{user_id}")
async def clear_cart_api(user_id: int):
    await clear_cart(user_id); return {"ok": True}

# ── ORDERS ────────────────────────────────────────────────────────────────────
class OrderRequest(BaseModel):
    user_id: int; address: str; latitude: float; longitude: float
    comment: Optional[str] = None
    coupon_code: Optional[str] = None

@app.post("/api/orders")
async def place_order(req: OrderRequest):
    cart_items = await get_cart(req.user_id)
    if not cart_items: raise HTTPException(400, "Cart is empty")

    subtotal = sum(i.product.price * i.quantity for i in cart_items)
    delivery_price, distance_km = calc_delivery(req.latitude, req.longitude)

    # Kupon
    discount_amount = 0
    coupon_code = None
    if req.coupon_code:
        coupon = await get_coupon(req.coupon_code)
        if coupon and subtotal >= coupon.min_order_amount:
            if coupon.discount_type == "percent":
                discount_amount = round(subtotal * coupon.discount_value / 100)
            else:
                discount_amount = min(coupon.discount_value, subtotal)
            coupon_code = req.coupon_code
            await use_coupon(req.coupon_code)

    order = await create_order(
        user_id=req.user_id, cart_items=cart_items,
        address=req.address, latitude=req.latitude, longitude=req.longitude,
        comment=req.comment, coupon_code=coupon_code,
        discount_amount=discount_amount, delivery_price=delivery_price,
        distance_km=distance_km
    )
    await clear_cart(req.user_id)

    # Guruhga yuborish (bot instance kerak — webhook orqali)
    # Bot handler dan chaqiriladi: handlers.send_order_to_group(bot, order, user)
    total_final = order.total_price + delivery_price - discount_amount
    return {
        "id": order.id, "total": order.total_price,
        "delivery_price": delivery_price, "discount": discount_amount,
        "total_final": total_final, "status": order.status
    }

@app.get("/api/orders/{user_id}")
async def user_orders(user_id: int):
    return [_order_dict(o) for o in await get_user_orders(user_id)]

@app.get("/api/admin/orders")
async def admin_orders(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    status: Optional[str] = None, _=Depends(verify_admin)
):
    df = datetime.fromisoformat(date_from) if date_from else None
    dt = datetime.fromisoformat(date_to + "T23:59:59") if date_to else None
    return [_order_dict(o) for o in await get_all_orders(df, dt, status)]

@app.put("/api/admin/orders/{order_id}/status")
async def change_order_status(order_id: int, status: str, _=Depends(verify_admin)):
    order = await get_order(order_id)
    if not order: raise HTTPException(404, "Order not found")
    await update_order_status(order_id, status)
    return {"ok": True}

# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────
@app.get("/api/admin/orders/export/excel")
async def export_orders_excel(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    status: Optional[str] = None, _=Depends(verify_admin)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl o'rnatilmagan: pip install openpyxl")

    df = datetime.fromisoformat(date_from) if date_from else None
    dt = datetime.fromisoformat(date_to + "T23:59:59") if date_to else None
    orders = await get_all_orders(df, dt, status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar"

    header_fill = PatternFill("solid", fgColor="5B5BD6")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    status_colors = {
        "pending": "FFF3CD", "confirmed": "D1ECF1",
        "delivering": "CCE5FF", "delivered": "D4EDDA", "cancelled": "F8D7DA"
    }

    headers = ["#", "Sana", "Mijoz", "Telefon", "Manzil", "Masofа (km)",
               "Mahsulotlar", "Mahsulot narxi", "Yetkazish", "Chegirma",
               "Jami", "Kupon", "Status", "Izoh"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 30
    col_widths = [6, 18, 20, 16, 35, 12, 45, 16, 14, 12, 16, 14, 16, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    status_uz = {"pending": "Kutilmoqda", "confirmed": "Tasdiqlandi",
                 "delivering": "Yetkazilmoqda", "delivered": "Yetkazildi", "cancelled": "Bekor"}

    for row_idx, o in enumerate(orders, 2):
        items_text = "; ".join([f"{i.product.title_uz if i.product else '?'} x{i.quantity}" for i in o.items])
        total_final = o.total_price + (o.delivery_price or 0) - (o.discount_amount or 0)
        row_data = [
            o.id,
            o.created_at.strftime("%d.%m.%Y %H:%M"),
            o.user.full_name if o.user else "—",
            o.user.phone if o.user else "—",
            o.address or "—",
            round(o.distance_km or 0, 1),
            items_text,
            f"{o.total_price:,.0f}",
            f"{o.delivery_price or 0:,.0f}",
            f"{o.discount_amount or 0:,.0f}",
            f"{total_final:,.0f}",
            o.coupon_code or "—",
            status_uz.get(o.status, o.status),
            o.comment or "—",
        ]
        fill_color = status_colors.get(o.status, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_color)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col in (8, 9, 10, 11): cell.alignment = Alignment(horizontal="right", vertical="center")

    # Jami
    total_sum = sum(o.total_price + (o.delivery_price or 0) - (o.discount_amount or 0) for o in orders)
    summary_row = len(orders) + 2
    ws.cell(summary_row, 1, "JAMI:").font = Font(bold=True)
    ws.cell(summary_row, 11, f"{total_sum:,.0f}").font = Font(bold=True, color="5B5BD6")
    ws.cell(summary_row, 13, f"Buyurtmalar: {len(orders)}").font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    fname = f"orders_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

# ── STATS ─────────────────────────────────────────────────────────────────────
@app.get("/api/admin/stats")
async def admin_stats(_=Depends(verify_admin)):
    return await get_stats()

# ── HELPERS ───────────────────────────────────────────────────────────────────
def _user_dict(u):
    return {"id": u.id, "telegram_id": u.telegram_id, "full_name": u.full_name,
            "username": u.username, "phone": u.phone, "language": u.language, "is_verified": u.is_verified}

def _cat_dict(c):
    return {"id": c.id, "name_uz": c.name_uz, "name_ru": c.name_ru,
            "image_url": c.image_url, "parent_id": c.parent_id, "is_active": c.is_active}

def _product_dict(p):
    return {"id": p.id, "title_uz": p.title_uz, "title_ru": p.title_ru,
            "description_uz": p.description_uz, "description_ru": p.description_ru,
            "price": p.price, "old_price": p.old_price, "images": p.images or [],
            "category_id": p.category_id, "stock": p.stock, "is_active": p.is_active}

def _order_dict(o):
    total_final = o.total_price + (o.delivery_price or 0) - (o.discount_amount or 0)
    return {
        "id": o.id, "status": o.status, "total_price": o.total_price,
        "delivery_price": o.delivery_price or 0, "discount_amount": o.discount_amount or 0,
        "total_final": total_final, "distance_km": o.distance_km,
        "address": o.address, "latitude": o.latitude, "longitude": o.longitude,
        "comment": o.comment, "coupon_code": o.coupon_code, "created_at": str(o.created_at),
        "user": _user_dict(o.user) if o.user else None,
        "items": [{"product_id": i.product_id, "quantity": i.quantity, "price": i.price,
                   "product_title": i.product.title_uz if i.product else "",
                   "product_img": (i.product.images[0] if i.product and i.product.images else None)}
                  for i in o.items]
    }
